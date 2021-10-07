from openpyxl import load_workbook, workbook
from models.food import Foods
workbook = load_workbook(filename='documents\ABBREV.xlsx')

array = []
sheet = workbook.active

data_frame = sheet['A:BA']
count = 1

for row in sheet.iter_rows():
    if count == 1:
        count += 1
        continue
    count += 1
    
    array.append({
        'NDB_No': row[0].value,
        'foodName': row[1].value,
        'water': row[2].value,
        'calorie': row[3].value,
        'protein': row[4].value,
        'lipid': row[5].value,
        'ash': row[6].value,
        'carbohydrate': row[7].value,
        'fiber': row[8].value,
        'sugar': row[9].value,
        'calcium': row[10].value,
        'iron': row[11].value,
        'magnesium': row[12].value,
        'phosphorus': row[13].value,
        'potassium': row[14].value,
        'sodium': row[15].value,
        'zinc': row[16].value,
        'copper': row[17].value,
        'manganese': row[18].value,
        'selenium': row[19].value,
        'vitaminC': row[20].value,
        'thiamin': row[21].value,
        'riboflavin': row[22].value,
        'niacin': row[23].value,
        'pantoAcid': row[24].value,
        'vitaminB6': row[25].value,
        'folate': row[26].value,
        'folicAcid': row[27].value,
        'foodFolate': row[28].value,
        'folateDFE': row[29].value,
        'choline': row[30].value,
        'vitaminB12': row[31].value,
        'vitaminAIU': row[32].value,
        'vitaminARAE': row[33].value,
        'retinol': row[34].value,
        'alphaCarot': row[35].value,
        'betaCarot': row[36].value,
        'betaCrypt': row[37].value,
        'lycopene': row[38].value,
        'lutANDzea': row[39].value,
        'vitaminE': row[40].value,
        'vitaminDUG': row[41].value,
        'vitaminDIU': row[42].value,
        'vitaminK': row[43].value,
        'faSat': row[44].value,
        'faMono': row[45].value,
        'faPoly': row[46].value,
        'cholestrl': row[47].value,
        'gmWt1': row[48].value,
        'gmWtDesc1': row[49].value,
        'gmWt2': row[50].value,
        'gmWtDesc2': row[51].value,
        'refusePct': row[52].value
        })


food_instances = [Foods(**data) for data in array]
Foods.objects.insert(food_instances, load_bulk = False)

print('USDA National Nutrient Data is copied successfully in your Database ;)')